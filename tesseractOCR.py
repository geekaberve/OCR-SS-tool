import pytesseract
from pytesseract import Output
import cv2
import logging
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
from PIL import Image, ImageDraw, ImageFont

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def initialize_tesseract():
    logger.info("Initializing Tesseract OCR...")
    # Set Tesseract command path if necessary (uncomment and set your path)
    # pytesseract.pytesseract.tesseract_cmd = r'/path/to/tesseract'
    return pytesseract

def process_image(file_path, ocr):
    try:
        # Load image
        image = cv2.imread(file_path)
        if image is None:
            raise ValueError("Could not open image!")

        logger.info("Processing image with Tesseract OCR...")

        # Convert to RGB
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # Run Tesseract OCR
        data = ocr.image_to_data(image_rgb, output_type=Output.DICT)

        n_boxes = len(data['level'])
        extracted_data = []

        for i in range(n_boxes):
            try:
                x = data['left'][i] + data['width'][i] / 2
                y = data['top'][i] + data['height'][i] / 2
                text = data['text'][i]
                confidence = float(data['conf'][i]) / 100.0 if data['conf'][i] != '-1' else 0.0
                bbox = [
                    (data['left'][i], data['top'][i]),
                    (data['left'][i] + data['width'][i], data['top'][i]),
                    (data['left'][i] + data['width'][i], data['top'][i] + data['height'][i]),
                    (data['left'][i], data['top'][i] + data['height'][i]),
                ]
                if text.strip():
                    extracted_data.append({'x': x, 'y': y, 'text': text.strip(), 'confidence': confidence, 'bbox': bbox})
            except Exception as e:
                logger.warning(f"Error processing box {i}: {str(e)}")

        return extracted_data

    except Exception as e:
        logger.error(f"Error processing image: {str(e)}")
        raise

def group_into_rows(data, y_threshold=10):
    # Sort data by y-coordinate
    data_sorted = sorted(data, key=lambda k: k['y'])

    # Group text into rows
    rows = []
    current_row = []
    last_y = None

    for item in data_sorted:
        x = item['x']
        y = item['y']
        text = item['text']
        if last_y is None or abs(y - last_y) > y_threshold:
            if current_row:
                rows.append(sorted(current_row, key=lambda k: k[0]))
            current_row = [(x, text, item['confidence'])]
            last_y = y
        else:
            current_row.append((x, text, item['confidence']))

    # Add the last row
    if current_row:
        rows.append(sorted(current_row, key=lambda k: k[0]))

    # Modify the return value to include confidence
    return [[(text, confidence) for x, text, confidence in row] for row in rows]

def save_as_xlsx(rows, output_xlsx, green_threshold=0.97, yellow_threshold=0.92):
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, row in enumerate(rows, start=1):
        for col_index, cell in enumerate(row, start=1):
            text, confidence = cell
            ws.cell(row=row_index, column=col_index, value=text)

            if confidence >= green_threshold:
                # Green for confidence >= green_threshold
                fill_color = '00FF00'
            elif confidence >= yellow_threshold:
                # Yellow for confidence >= yellow_threshold
                fill_color = 'FFFF00'
            else:
                # Red for confidence below yellow_threshold
                fill_color = 'FF0000'

            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.cell(row=row_index, column=col_index).fill = fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_xlsx)
    logger.info(f"Excel file has been saved at: {output_xlsx}")

def draw_bounding_boxes(image_path, data, output_path):
    try:
        image = cv2.imread(image_path)
        for item in data:
            bbox = item['bbox']
            if len(bbox) != 4:
                logger.warning(f"Skipping invalid bounding box: {bbox}")
                continue
            
            try:
                # Convert bbox to numpy array of integers
                bbox_np = np.array(bbox, dtype=np.int32)
                
                # Ensure the bbox is in the correct format for cv2.polylines
                bbox_np = bbox_np.reshape((-1, 1, 2))
                
                cv2.polylines(image, [bbox_np], True, (0, 255, 0), 2)
                
                # Add text above the bounding box in black color
                text = item['text']
                confidence = item['confidence']
                label = f"{text} ({confidence:.2f})"
                cv2.putText(image, label, (bbox[0][0], bbox[0][1] - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 2)
            except Exception as e:
                logger.warning(f"Error drawing bounding box: {str(e)}")
        
        cv2.imwrite(output_path, image)
        logger.info(f"Image with bounding boxes saved to {output_path}")
    except Exception as e:
        logger.error(f"Error in draw_bounding_boxes: {str(e)}")
        raise
