import easyocr
import cv2
import logging
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
from PIL import Image, ImageDraw, ImageFont

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def initialize_easyocr():
    logger.info("Initializing EasyOCR...")
    # Initialize the reader with desired languages
    reader = easyocr.Reader(['en'])  # You can specify other languages if needed
    return reader

def process_image(file_path, reader):
    try:
        # Read image
        image = cv2.imread(file_path)
        if image is None:
            raise ValueError("Could not open image!")

        logger.info("Processing image with EasyOCR...")

        # Convert to RGB
        image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # Perform OCR
        result = reader.readtext(image_rgb)

        extracted_data = []

        for res in result:
            bbox, text, confidence = res
            # Calculate the center point of the bounding box
            x = sum([point[0] for point in bbox]) / 4.0
            y = sum([point[1] for point in bbox]) / 4.0
            extracted_data.append({'x': x, 'y': y, 'text': text.strip(), 'confidence': confidence, 'bbox': bbox})

        return extracted_data

    except Exception as e:
        logger.error(f"Error processing image: {str(e)}")
        raise

def group_into_rows(data, y_threshold=10):
    # Sort data by y-coordinate
    data_sorted = sorted(data, key=lambda k: k['y'])

    # Group text into rows
    rows = []
    current_row = []
    last_y = None

    for item in data_sorted:
        x = item['x']
        y = item['y']
        text = item['text']
        confidence = item['confidence']
        if last_y is None or abs(y - last_y) > y_threshold:
            if current_row:
                rows.append(sorted(current_row, key=lambda k: k[0]))
            current_row = [(x, text, confidence)]
            last_y = y
        else:
            current_row.append((x, text, confidence))

    # Add the last row
    if current_row:
        rows.append(sorted(current_row, key=lambda k: k[0]))

    # Return rows with text and confidence
    return [[(text, confidence) for x, text, confidence in row] for row in rows]

def save_as_xlsx(rows, output_xlsx, green_threshold=0.97, yellow_threshold=0.92):
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, row in enumerate(rows, start=1):
        for col_index, cell in enumerate(row, start=1):
            text, confidence = cell
            ws.cell(row=row_index, column=col_index, value=text)

            if confidence >= green_threshold:
                # Green for confidence >= green_threshold
                fill_color = '00FF00'
            elif confidence >= yellow_threshold:
                # Yellow for confidence >= yellow_threshold
                fill_color = 'FFFF00'
            else:
                # Red for confidence below yellow_threshold
                fill_color = 'FF0000'

            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.cell(row=row_index, column=col_index).fill = fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_xlsx)
    logger.info(f"Excel file has been saved at: {output_xlsx}")

def draw_bounding_boxes(image_path, data, output_image_path):
    image = Image.open(image_path)
    # Convert image to RGB mode
    image = image.convert('RGB')
    draw = ImageDraw.Draw(image)
    try:
        font = ImageFont.truetype("arial.ttf", 16)  # Use a true type font
    except:
        font = ImageFont.load_default()

    for item in data:
        bbox = item['bbox']
        text = item['text']
        confidence = item['confidence']

        # Convert bbox coordinates to integer tuples
        bbox = [(int(point[0]), int(point[1])) for point in bbox]

        # Draw bounding box
        draw.polygon(bbox, outline='green')

        # Put text and confidence
        x, y = bbox[0][0], bbox[0][1]
        draw.text((x, y - 20), f'{text} ({confidence:.2f})', fill='red', font=font)

    image.save(output_image_path)
    logger.info(f"Image with bounding boxes saved at: {output_image_path}")
