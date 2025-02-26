import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os
import threading
from OCR_Modules.paddleOCR import initialize_ocr_SLANet_LCNetV2, process_image as paddle_process_image, group_into_rows as paddle_group_into_rows, save_as_xlsx as paddle_save_as_xlsx, draw_bounding_boxes as paddle_draw_bounding_boxes
from OCR_Modules.tesseractOCR import initialize_tesseract, process_image as tesseract_process_image, group_into_rows as tesseract_group_into_rows, save_as_xlsx as tesseract_save_as_xlsx, draw_bounding_boxes as tesseract_draw_bounding_boxes
import tempfile
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import subprocess
import sys
from PIL import Image, ImageGrab
import keyboard
import win32clipboard
from io import BytesIO
import time
from tkinter import messagebox
import shutil
from datetime import datetime
import psutil
import win32api
import win32con

def get_resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        # Running in PyInstaller bundle
        base_path = sys._MEIPASS
    else:
        # Running in normal Python environment
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class OCRApp:
    def __init__(self, root):
        # Get the application's installation directory
        if getattr(sys, 'frozen', False):
            self.app_dir = os.path.dirname(sys.executable)
            # Force PyInstaller to use the installation directory
            if hasattr(sys, '_MEIPASS'):
                sys._MEIPASS = self.app_dir
        else:
            self.app_dir = os.path.dirname(os.path.abspath(__file__))
            
        # Set up environment variables for dependencies
        os.environ['TESSDATA_PREFIX'] = os.path.join(self.app_dir, 'tessdata')
        os.environ['PATH'] = os.path.join(self.app_dir, 'tesseract_binary') + os.pathsep + os.environ.get('PATH', '')
        
        # Force PaddleOCR to use the installation directory
        os.environ['PADDLE_OCR_PATH'] = os.path.join(self.app_dir, 'paddleocr')
        os.environ['PADDLE_PATH'] = os.path.join(self.app_dir, 'paddle')
        
        # Add dependency directories to Python path
        paddle_dir = os.path.join(self.app_dir, 'paddle')
        paddleocr_dir = os.path.join(self.app_dir, 'paddleocr')
        
        # Ensure these directories are at the start of sys.path
        sys.path.insert(0, paddle_dir)
        sys.path.insert(0, paddleocr_dir)
        
        # Initialize the rest of the application
        self.initialize_app(root)

    def initialize_app(self, root):
        self.root = root
        self.root.title("OCR to Excel Tool")
        self.root.geometry("1280x720")
        
        # Use absolute paths for all resources
        icon_path = get_resource_path('icons/icon.png')
        self.icon_image = tk.PhotoImage(file=icon_path)
        self.root.iconphoto(False, self.icon_image)
        
        self.style = ttk.Style('cosmo')
        self.ocr_engine = tk.StringVar()
        self.ocr_engine.set("PaddleOCR")

        self.green_threshold = tk.IntVar(value=97)
        self.yellow_threshold = tk.IntVar(value=92)
        self.output_directory = None
        self.is_screenshot = False
        
        self.setup_ui()

    def setup_ui(self):
        # Main frame
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Center frame for initial view
        self.center_frame = ttk.Frame(self.main_frame)
        self.center_frame.pack(expand=True)

        # Logo
        logo_img = Image.open("icons/icon.png")
        logo_img = logo_img.resize((100, 100), Image.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(logo_img)
        self.logo_label = ttk.Label(self.center_frame, image=self.logo_photo)
        self.logo_label.pack(pady=(20, 10))

        # Output Directory Selection
        folder_icon = Image.open("icons/folder.png")
        folder_icon = folder_icon.resize((20, 20), Image.LANCZOS)
        self.folder_icon_photo = ImageTk.PhotoImage(folder_icon)
        output_dir_button = ttk.Button(self.center_frame, text="Select Output Directory", 
                                       image=self.folder_icon_photo, compound=tk.LEFT,
                                       command=self.select_output_directory)
        output_dir_button.pack(pady=(10, 5))

        # OCR Engine Selection
        ocr_label = ttk.Label(self.center_frame, text="Select OCR Engine:")
        ocr_label.pack(pady=(10, 5))
        ocr_dropdown = ttk.Combobox(self.center_frame, textvariable=self.ocr_engine, state="readonly", width=30)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract')
        ocr_dropdown.pack(pady=(0, 20))

        # Confidence Thresholds
        thresholds_frame = ttk.Frame(self.center_frame)
        thresholds_frame.pack(pady=(10, 20))

        options = [str(i) for i in range(80, 101)]  # Values from 80 to 100

        green_label = ttk.Label(thresholds_frame, text="High Confidence Threshold (Green) (> %):")
        green_label.grid(row=0, column=0, padx=5, pady=5, sticky='e')
        green_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.green_threshold, values=options, state='readonly', width=5)
        green_dropdown.grid(row=0, column=1, padx=5, pady=5)
        green_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        yellow_label = ttk.Label(thresholds_frame, text="Medium Confidence Threshold (Yellow) (> %):")
        yellow_label.grid(row=1, column=0, padx=5, pady=5, sticky='e')
        yellow_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.yellow_threshold, values=options, state='readonly', width=5)
        yellow_dropdown.grid(row=1, column=1, padx=5, pady=5)
        yellow_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        # Upload Button
        upload_icon = Image.open("icons/upload.png")
        upload_icon = upload_icon.resize((20, 20), Image.LANCZOS)
        self.upload_icon_photo = ImageTk.PhotoImage(upload_icon)
        self.upload_button = ttk.Button(self.center_frame, text="Upload Image", 
                                        image=self.upload_icon_photo, compound=tk.LEFT,
                                        command=self.select_image, width=20)
        self.upload_button.pack(pady=(0, 10))

        # Screenshot Button
        screenshot_icon = Image.open("icons/screenshot.png")
        screenshot_icon = screenshot_icon.resize((20, 20), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        self.screenshot_button = ttk.Button(self.center_frame, text="Screenshot", 
                                            image=self.screenshot_icon_photo, compound=tk.LEFT,
                                            command=self.take_screenshot, width=20)
        self.screenshot_button.pack(pady=(0, 20))

        # Status Label
        self.status_label = ttk.Label(self.center_frame, text="")
        self.status_label.pack(pady=(0, 10))

        # Progress bar (hidden initially)
        self.progress_bar = ttk.Progressbar(self.center_frame, mode='indeterminate')
        self.progress_bar.pack(pady=(0, 10))
        self.progress_bar.pack_forget()

        # Prepare frames for results
        self.top_frame = ttk.Frame(self.main_frame)
        self.left_frame = ttk.Frame(self.main_frame)
        self.middle_frame = ttk.Frame(self.main_frame)
        self.right_frame = ttk.Frame(self.main_frame)  # For the sidebar

    def select_output_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_directory = directory
            self.status_label.config(text=f"Output directory set to: {self.output_directory}")

    def update_thresholds(self, event=None):
        # Ensure that green_threshold is always greater than yellow_threshold
        if self.green_threshold.get() < self.yellow_threshold.get():
            self.yellow_threshold.set(self.green_threshold.get())

    def reset_ui(self):
        # Hide the top, left, middle, and right frames
        self.top_frame.pack_forget()
        self.left_frame.pack_forget()
        self.middle_frame.pack_forget()
        self.right_frame.pack_forget()

        # Clear images from frames
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Show the center frame
        self.center_frame.pack(expand=True)

        # Reset status label
        self.status_label.config(text="")

        # Remove progress bar
        self.progress_bar.pack_forget()

    def select_image(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Image files", ("*.png", "*.jpg", "*.jpeg", "*.bmp", "*.tiff"))]
        )
        if file_path:
            self.reset_ui()  # Reset the UI before processing a new image
            self.is_screenshot = False
            self.process_image(file_path)

    def take_screenshot(self):
        # Minimize the root window
        self.root.withdraw()
        # Capture the screenshot
        image = self.capture_screenshot()
        
        # Only process if we got a screenshot
        if image:
            # Restore the root window
            self.root.deiconify()
            
            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                # For screenshots, default to Desktop
                output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            # Save the screenshot image
            base_filename = "screenshot"
            date_string = time.strftime("%Y-%m-%d_%H-%M-%S")
            screenshot_path = os.path.join(output_dir, base_filename + "_" + date_string + ".png")
            self.is_screenshot = True
            image.save(screenshot_path)
            # Reset the UI before processing
            self.reset_ui()
            # Process the image
            self.process_image(screenshot_path)

    def capture_screenshot(self):
        
        from PIL import ImageGrab        
        try:
            # Store current window state
            window_state = self.root.state()
            
            # Minimize the window
            self.root.withdraw()

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
            
            # Simulate Windows + Shift + S keystroke
            keyboard.press_and_release('windows+shift+s')
            start_time = time.time()
            # Wait for 10 seconds
            timeout = 30

            while True:
                # Try to get the screenshot from clipboard
                screenshot = ImageGrab.grabclipboard() 
                if screenshot is not None:
                    return screenshot
                else:
                    # Restore window before showing error
                    if time.time() - start_time > timeout:
                        self.root.deiconify()
                        messagebox.showerror("Screenshot Error", 
                        "No screenshot was taken within the time limit (30 seconds).\n"
                        "Please try again and make sure to complete the screenshot within the time limit.")
                        raise TimeoutError("Retry taking screenshot within 30 seconds")
                    time.sleep(2)
        except TimeoutError as e:
            raise e
        except Exception as e:
            # Always restore window in case of any error
            self.root.deiconify()
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            raise Exception(f"Error capturing screenshot: {str(e)}")

    def process_image(self, file_path):
        # Start a new thread for processing
        processing_thread = threading.Thread(target=self._process_image_thread, args=(file_path,))
        processing_thread.start()

    def _process_image_thread(self, file_path):
        try:
            ocr_engine = self.ocr_engine.get()

            # Show progress bar in the center frame
            self.progress_bar.pack(pady=(0, 10))
            self.progress_bar.start()

            if ocr_engine == "PaddleOCR":
                self.process_with_paddleocr(file_path)
            elif ocr_engine == "Tesseract":
                self.process_with_tesseract(file_path)
            else:
                raise ValueError("Please select an OCR engine.")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")
        finally:
            # Hide progress bar
            self.progress_bar.stop()
            self.progress_bar.pack_forget()

    def process_with_paddleocr(self, file_path):
        try:
            # Initialize OCR with explicit paths
            ocr = initialize_ocr_SLANet_LCNetV2(
                model_dir=os.path.join(self.app_dir, 'paddleocr', 'whl')
            )
            
            data = paddle_process_image(file_path, ocr)
            
            rows = paddle_group_into_rows(data)
            
            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    # For screenshots, default to Desktop
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    # For uploaded images, use the same directory as the image
                    output_dir = os.path.dirname(file_path)
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            # Create the output filenames
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            # Get thresholds from dropdowns
            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            paddle_save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)

            paddle_draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")

    def process_with_tesseract(self, file_path):
        try:
            # Initialize Tesseract with explicit paths
            ocr = initialize_tesseract(
                tesseract_cmd=os.path.join(self.app_dir, 'tesseract_binary', 'tesseract.exe')
            )
            
            data = tesseract_process_image(file_path, ocr)

            if not data:
                raise ValueError("No data extracted from image.")

            rows = tesseract_group_into_rows(data)

            if not rows:
                raise ValueError("No rows extracted from data.")

            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    # For screenshots, default to Desktop
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    # For uploaded images, use the same directory as the image
                    output_dir = os.path.dirname(file_path)
            # Ensure output directory exists
            os.makedirs(output_dir, exist_ok=True)
            # Create the output filenames
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            # Get thresholds from dropdowns
            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            tesseract_save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)

            tesseract_draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except ValueError as ve:
            self.status_label.config(text=f"Error: {str(ve)}\nPlease try a different image or OCR engine.")
        except Exception as e:
            self.status_label.config(text=f"Unexpected error: {str(e)}\nPlease try a different image or OCR engine.")

    def display_results(self, image_path, excel_path):
        self.reorganize_layout()

        # Clear previous images
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Display image with bounding boxes
        self.display_image(image_path, self.left_frame)

        # Display Excel image
        excel_image_path = os.path.splitext(excel_path)[0] + "_excel_image.png"
        self.generate_excel_image(excel_path, excel_image_path)

        # Add padding to the middle frame
        padding_frame = ttk.Frame(self.middle_frame, padding=20)
        padding_frame.pack(fill=tk.BOTH, expand=True)
        self.display_image(excel_image_path, padding_frame)

        # Setup the sidebar
        self.setup_sidebar()

    def reorganize_layout(self):
        # Hide the center frame
        self.center_frame.pack_forget()

        # Clear the top_frame
        for widget in self.top_frame.winfo_children():
            widget.destroy()

        # Set up top frame
        self.top_frame.pack(side=tk.TOP, fill=tk.X, pady=(10, 0))

        # Create a frame inside top_frame to center widgets
        top_inner_frame = ttk.Frame(self.top_frame)
        top_inner_frame.pack(anchor='center')

        # Add small logo
        small_logo_img = Image.open("icons/icon.png")
        small_logo_img = small_logo_img.resize((50, 50), Image.LANCZOS)
        self.small_logo_photo = ImageTk.PhotoImage(small_logo_img)
        logo_label = ttk.Label(top_inner_frame, image=self.small_logo_photo)
        logo_label.pack(side=tk.LEFT, padx=(10, 5))

        ocr_label = ttk.Label(top_inner_frame, text="Select OCR Engine:")
        ocr_label.pack(side=tk.LEFT, padx=(10, 5))
        ocr_dropdown = ttk.Combobox(top_inner_frame, textvariable=self.ocr_engine, state="readonly", width=20)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract')
        ocr_dropdown.pack(side=tk.LEFT, padx=(0, 10))
        upload_button = ttk.Button(top_inner_frame, text="Upload Image", command=self.select_image)
        upload_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Screenshot Button with Icon
        screenshot_icon = Image.open(os.path.join('icons', 'screenshoticon.png'))
        screenshot_icon = screenshot_icon.resize((30, 30), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        screenshot_button = ttk.Button(top_inner_frame, image=self.screenshot_icon_photo, command=self.take_screenshot)
        screenshot_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Open Folder Button with Icon
        folder_icon = Image.open(os.path.join('icons', 'folder.png'))
        folder_icon = folder_icon.resize((30, 30), Image.LANCZOS)
        self.folder_icon_photo = ImageTk.PhotoImage(folder_icon)
        folder_button = ttk.Button(top_inner_frame, image=self.folder_icon_photo, command=self.open_output_folder)
        folder_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Home Button with Icon
        home_icon = Image.open(os.path.join('icons', 'home.png'))
        home_icon = home_icon.resize((30, 30), Image.LANCZOS)
        self.home_icon_photo = ImageTk.PhotoImage(home_icon)
        home_button = ttk.Button(top_inner_frame, image=self.home_icon_photo, command=self.reset_ui)
        home_button.pack(side=tk.LEFT, padx=(0, 10))

        # Set up frames
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)

    def display_image(self, image_path, panel):
        image = Image.open(image_path)

        # Create a canvas to display the image
        canvas = tk.Canvas(panel, bg='white')
        canvas.pack(fill=tk.BOTH, expand=True)

        # Function to resize the image when the panel size changes
        def resize_image(event):
            # Get the size of the canvas
            canvas_width = event.width
            canvas_height = event.height

            # Calculate the scaling factor
            width_ratio = canvas_width / image.width
            height_ratio = canvas_height / image.height
            scale_factor = min(width_ratio, height_ratio, 1.0)  # Do not upscale images

            new_width = int(image.width * scale_factor)
            new_height = int(image.height * scale_factor)

            resized_image = image.resize((new_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized_image)

            canvas.delete("all")
            canvas.create_image(canvas_width/2, canvas_height/2, image=photo, anchor='center')
            canvas.image = photo  # Keep a reference

        # Bind the resize event to the function
        canvas.bind("<Configure>", resize_image)

    def generate_excel_image(self, excel_path, output_image_path):
        from openpyxl import load_workbook
        from PIL import Image, ImageDraw, ImageFont

        wb = load_workbook(excel_path)
        ws = wb.active

        # Increase cell size and font size
        cell_width = 90
        cell_height = 30
        font_size = 20
        border_size = 40  # Increased border size

        max_col = ws.max_column
        max_row = ws.max_row
        image_width = cell_width * max_col + 2 * border_size
        image_height = cell_height * max_row + 2 * border_size

        image = Image.new('RGB', (image_width, image_height), 'white')
        draw = ImageDraw.Draw(image)
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()

        for row in ws.iter_rows():
            for cell in row:
                col_idx = cell.column - 1
                row_idx = cell.row - 1
                x1 = col_idx * cell_width + border_size
                y1 = row_idx * cell_height + border_size
                x2 = x1 + cell_width
                y2 = y1 + cell_height

                # Draw cell background
                fill_color = 'FFFFFF'  # Default fill
                if cell.fill and cell.fill.fgColor:
                    if cell.fill.fgColor.type == 'rgb':
                        fill_color = cell.fill.fgColor.rgb[-6:]
                    elif cell.fill.fgColor.type == 'indexed':
                        fill_color = 'FFFFFF'  # Handle indexed colors as white

                draw.rectangle([x1, y1, x2, y2], fill=f'#{fill_color}', outline='black')

                # Draw cell text
                text = str(cell.value) if cell.value is not None else ''
                bbox = font.getbbox(text)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                text_x = x1 + (cell_width - text_width) / 2
                text_y = y1 + (cell_height - text_height) / 2
                draw.text((text_x, text_y), text, fill='black', font=font)

        image.save(output_image_path)

    def setup_sidebar(self):
        # Sidebar content
        sidebar_frame = ttk.Frame(self.right_frame, padding=10)
        sidebar_frame.pack(fill=tk.Y, expand=False)

        # Accuracy Box
        accuracy_label = ttk.Label(sidebar_frame, text="Accuracy Thresholds", font=('Helvetica', 14, 'bold'))
        accuracy_label.pack(pady=(0, 10))

        # Icons for accuracy levels
        green_icon = Image.open(os.path.join('icons', 'green_circle.png'))
        green_icon = green_icon.resize((20, 20), Image.LANCZOS)
        self.green_icon_photo = ImageTk.PhotoImage(green_icon)

        yellow_icon = Image.open(os.path.join('icons', 'yellow_circle.png'))
        yellow_icon = yellow_icon.resize((20, 20), Image.LANCZOS)
        self.yellow_icon_photo = ImageTk.PhotoImage(yellow_icon)

        red_icon = Image.open(os.path.join('icons', 'red_circle.png'))
        red_icon = red_icon.resize((20, 20), Image.LANCZOS)
        self.red_icon_photo = ImageTk.PhotoImage(red_icon)

        # Green Threshold Label
        green_frame = ttk.Frame(sidebar_frame)
        green_frame.pack(pady=5, anchor='w')
        green_label_icon = ttk.Label(green_frame, image=self.green_icon_photo)
        green_label_icon.pack(side=tk.LEFT)
        green_label_text = ttk.Label(green_frame, text=f"High Confidence (> {self.green_threshold.get()}%)")
        green_label_text.pack(side=tk.LEFT, padx=5)

        # Yellow Threshold Label
        yellow_frame = ttk.Frame(sidebar_frame)
        yellow_frame.pack(pady=5, anchor='w')
        yellow_label_icon = ttk.Label(yellow_frame, image=self.yellow_icon_photo)
        yellow_label_icon.pack(side=tk.LEFT)
        yellow_label_text = ttk.Label(yellow_frame, text=f"Medium Confidence (> {self.yellow_threshold.get()}%)")
        yellow_label_text.pack(side=tk.LEFT, padx=5)

        # Red Threshold Label
        red_frame = ttk.Frame(sidebar_frame)
        red_frame.pack(pady=5, anchor='w')
        red_label_icon = ttk.Label(red_frame, image=self.red_icon_photo)
        red_label_icon.pack(side=tk.LEFT)
        red_label_text = ttk.Label(red_frame, text=f"Low Confidence (≤ {self.yellow_threshold.get()}%)")
        red_label_text.pack(side=tk.LEFT, padx=5)

        # Divider
        separator = ttk.Separator(sidebar_frame, orient='horizontal')
        separator.pack(fill='x', pady=10)

        # Disclaimer Box
        disclaimer_frame = ttk.Frame(sidebar_frame)
        disclaimer_frame.pack(pady=(10, 10))

        # Info Icon
        info_icon = Image.open(os.path.join('icons', 'info.png'))
        info_icon = info_icon.resize((20, 20), Image.LANCZOS)
        self.info_icon_photo = ImageTk.PhotoImage(info_icon)

        disclaimer_title_frame = ttk.Frame(disclaimer_frame)
        disclaimer_title_frame.pack(anchor='w')

        disclaimer_icon_label = ttk.Label(disclaimer_title_frame, image=self.info_icon_photo)
        disclaimer_icon_label.pack(side=tk.LEFT)

        disclaimer_label = ttk.Label(disclaimer_title_frame, text="Disclaimer", font=('Helvetica', 14, 'bold'))
        disclaimer_label.pack(side=tk.LEFT, padx=5)

        # Enhanced Disclaimer Text
        disclaimer_text = (
            "Note: When the input image contains a table with missing data in some rows or columns, "
            "the extracted data may not align perfectly in the output Excel file. "
            "Empty cells might cause subsequent data to shift positions, resulting in misaligned columns. "
            "Please review the Excel output carefully and adjust as needed."
        )
        disclaimer_message = ttk.Label(disclaimer_frame, text=disclaimer_text, wraplength=250, justify='left')
        disclaimer_message.pack(pady=5)

    def open_output_folder(self):
        """Open the output directory in the system's file explorer"""
        if self.output_directory:
            output_dir = self.output_directory
        else:
            # If no output directory was selected, use Desktop for screenshots or the image's directory
            if hasattr(self, 'is_screenshot') and self.is_screenshot:
                output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            else:
                output_dir = os.path.dirname(self.current_image_path) if hasattr(self, 'current_image_path') else None

        if output_dir and os.path.exists(output_dir):
            if sys.platform == 'darwin':  # macOS
                subprocess.run(['open', output_dir])
            elif sys.platform == 'win32':  # Windows
                subprocess.run(['explorer', output_dir])
            else:  # Linux
                subprocess.run(['xdg-open', output_dir])
        else:
            self.status_label.config(text="Output directory not found")

if __name__ == "__main__":
    root = ttk.Window(themename="cosmo")
    app = OCRApp(root)
    root.mainloop()