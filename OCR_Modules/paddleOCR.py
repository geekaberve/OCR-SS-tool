from paddleocr import PaddleOCR
import cv2
import logging
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os  # Added to use os.cpu_count()
from PIL import Image, ImageDraw, ImageFont
import sys
import traceback

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def initialize_ocr_SLANet_LCNetV2(model_dir=None):
    if model_dir is None:
        if getattr(sys, 'frozen', False):
            app_dir = os.path.dirname(sys.executable)
            model_dir = os.path.join(app_dir, 'paddleocr', 'whl')
        else:
            model_dir = os.path.expanduser('~/.paddleocr/whl')
    
    return PaddleOCR(
        use_angle_cls=True,
        lang='en',
        use_gpu=False,
        show_log=False,
        det_model_dir=os.path.join(model_dir, 'det'),
        cls_model_dir=os.path.join(model_dir, 'cls'),
        rec_model_dir=os.path.join(model_dir, 'rec')
    )

def process_image(file_path, ocr):
    try:
        # Load image
        logger.info(f"Loading image from: {file_path}")
        image = cv2.imread(file_path)
        if image is None:
            raise ValueError("Could not open image!")

        logger.info("Processing image...")

        # Perform OCR with error checking
        result = ocr.ocr(image, cls=True, det=True)
        logger.info(f"OCR result type: {type(result)}")
        logger.info(f"OCR result: {result}")

        if result is None:
            raise ValueError("OCR processing failed - no results returned")

        if not result:
            raise ValueError("No text detected in image.")

        # Extract text, coordinates, and confidence
        data = []
        for line in result:
            if not line:
                continue
            for word_info in line:
                if not word_info or len(word_info) != 2:  # Validate word_info structure
                    logger.warning(f"Skipping invalid word_info: {word_info}")
                    continue
                    
                bbox, (text, confidence) = word_info
                if not bbox or len(bbox) != 4:  # Validate bbox
                    logger.warning(f"Skipping invalid bbox: {bbox}")
                    continue
                    
                x = (bbox[0][0] + bbox[2][0]) / 2  # Average x-coordinate
                y = (bbox[0][1] + bbox[2][1]) / 2  # Average y-coordinate
                data.append({
                    'x': x, 
                    'y': y, 
                    'text': text, 
                    'confidence': confidence, 
                    'bbox': bbox
                })

        if not data:
            raise ValueError("No valid data extracted from OCR results")

        return data

    except Exception as e:
        error_msg = f"Error: {str(e)}\n\nTraceback:\n{traceback.format_exc()}"
        raise Exception(error_msg)

def group_into_rows(data, y_threshold=10):
    # Sort data by y-coordinate
    data_sorted = sorted(data, key=lambda k: k['y'])

    # Group text into rows
    rows = []
    current_row = []
    last_y = None

    for item in data_sorted:
        x = item['x']
        y = item['y']
        text = item['text']
        if last_y is None or abs(y - last_y) > y_threshold:
            if current_row:
                rows.append(sorted(current_row, key=lambda k: k[0]))
            current_row = [(x, text, item['confidence'])]
            last_y = y
        else:
            current_row.append((x, text, item['confidence']))

    # Add the last row
    if current_row:
        rows.append(sorted(current_row, key=lambda k: k[0]))

    # Modify the return value to include confidence
    return [[(text, confidence) for x, text, confidence in row] for row in rows]

def save_as_xlsx(rows, output_xlsx, green_threshold=0.97, yellow_threshold=0.92):
    wb = openpyxl.Workbook()
    ws = wb.active

    for row_index, row in enumerate(rows, start=1):
        for col_index, cell in enumerate(row, start=1):
            text, confidence = cell
            ws.cell(row=row_index, column=col_index, value=text)

            if confidence >= green_threshold:
                # Green for confidence >= green_threshold
                fill_color = '00FF00'
            elif confidence >= yellow_threshold:
                # Yellow for confidence >= yellow_threshold
                fill_color = 'FFFF00'
            else:
                # Red for confidence below yellow_threshold
                fill_color = 'FF0000'

            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
            ws.cell(row=row_index, column=col_index).fill = fill

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_xlsx)
    logger.info(f"Excel file has been saved at: {output_xlsx}")

def draw_bounding_boxes(image_path, data, output_image_path):
    image = Image.open(image_path)
    # Convert image to RGB mode
    image = image.convert('RGB')
    draw = ImageDraw.Draw(image)
    try:
        font = ImageFont.truetype("arial.ttf", 16)  # Use a true type font
    except:
        font = ImageFont.load_default()

    for item in data:
        bbox = item['bbox']
        text = item['text']
        confidence = item['confidence']

        # Draw bounding box
        bbox_points = [(point[0], point[1]) for point in bbox]
        draw.line(bbox_points + [bbox_points[0]], fill='green', width=2)

        # Put text and confidence
        x, y = bbox[0][0], bbox[0][1]
        draw.text((x, y - 20), f'{text} ({confidence:.2f})', fill='red', font=font)

    image.save(output_image_path)
    logger.info(f"Image with bounding boxes saved at: {output_image_path}")

if __name__ == "__main__":
    ocr_model = initialize_ocr_SLANet_LCNetV2()
    logger.info("PaddleOCR initialized successfully.")

    # Example usage
    image_path = 'example_image.jpg'
    output_xlsx = 'output.xlsx'
    output_image_path = 'output_image.jpg'

    ocr_data = process_image(image_path, ocr_model)
    rows = group_into_rows(ocr_data)
    save_as_xlsx(rows, output_xlsx)
    draw_bounding_boxes(image_path, ocr_data, output_image_path)
