import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk, ImageDraw, ImageFont, ImageGrab
import os
import threading
import tempfile
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import subprocess
import sys
import keyboard
import win32clipboard
from io import BytesIO
import time
from tkinter import messagebox
import shutil
from datetime import datetime
import psutil
import win32api
import win32con
import logging
import cv2
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from paddleocr import PaddleOCR
import pytesseract
from pytesseract import Output

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class PaddleOCREngine:
    def __init__(self, model_dir=None):
        if model_dir is None:
            model_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models')
        
        self.ocr = PaddleOCR(
            use_angle_cls=True,
            lang='en',
            use_gpu=False,
            show_log=False,
            det_model_dir=os.path.join(model_dir, 'det'),
            cls_model_dir=os.path.join(model_dir, 'cls'),
            rec_model_dir=os.path.join(model_dir, 'rec')
        )

    def process_image(self, file_path):
        try:
            image = cv2.imread(file_path)
            if image is None:
                raise ValueError("Could not open image!")

            result = self.ocr.ocr(image, cls=True, det=True)
            if result is None or not result:
                raise ValueError("No text detected in image.")

            data = []
            for line in result:
                if not line:
                    continue
                for word_info in line:
                    if not word_info or len(word_info) != 2:
                        continue
                    bbox, (text, confidence) = word_info
                    if not bbox or len(bbox) != 4:
                        continue
                    x = (bbox[0][0] + bbox[2][0]) / 2
                    y = (bbox[0][1] + bbox[2][1]) / 2
                    data.append({
                        'x': x, 
                        'y': y, 
                        'text': text, 
                        'confidence': confidence, 
                        'bbox': bbox
                    })

            return data
        except Exception as e:
            logger.error(f"Error processing image: {str(e)}")
            raise

    def group_into_rows(self, data, y_threshold=10):
        data_sorted = sorted(data, key=lambda k: k['y'])
        rows = []
        current_row = []
        last_y = None

        for item in data_sorted:
            x = item['x']
            y = item['y']
            text = item['text']
            if last_y is None or abs(y - last_y) > y_threshold:
                if current_row:
                    rows.append(sorted(current_row, key=lambda k: k[0]))
                current_row = [(x, text, item['confidence'])]
                last_y = y
            else:
                current_row.append((x, text, item['confidence']))

        if current_row:
            rows.append(sorted(current_row, key=lambda k: k[0]))

        return [[(text, confidence) for x, text, confidence in row] for row in rows]

    def save_as_xlsx(self, rows, output_xlsx, green_threshold=0.97, yellow_threshold=0.92):
        wb = openpyxl.Workbook()
        ws = wb.active

        for row_index, row in enumerate(rows, start=1):
            for col_index, cell in enumerate(row, start=1):
                text, confidence = cell
                ws.cell(row=row_index, column=col_index, value=text)

                if confidence >= green_threshold:
                    fill_color = '00FF00'
                elif confidence >= yellow_threshold:
                    fill_color = 'FFFF00'
                else:
                    fill_color = 'FF0000'

                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                ws.cell(row=row_index, column=col_index).fill = fill

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_xlsx)

    def draw_bounding_boxes(self, image_path, data, output_image_path):
        image = Image.open(image_path)
        image = image.convert('RGB')
        draw = ImageDraw.Draw(image)
        try:
            font = ImageFont.truetype("arial.ttf", 16)
        except:
            font = ImageFont.load_default()

        for item in data:
            bbox = item['bbox']
            text = item['text']
            confidence = item['confidence']

            bbox_points = [(point[0], point[1]) for point in bbox]
            draw.line(bbox_points + [bbox_points[0]], fill='green', width=2)
            x, y = bbox[0][0], bbox[0][1]
            draw.text((x, y - 20), f'{text} ({confidence:.2f})', fill='red', font=font)

        image.save(output_image_path)

class TesseractOCREngine:
    def __init__(self, tesseract_cmd=None):
        if tesseract_cmd is None:
            tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        
        pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
        os.environ['TESSDATA_PREFIX'] = r'C:\Program Files\Tesseract-OCR\tessdata'

    def process_image(self, file_path):
        try:
            image = cv2.imread(file_path)
            if image is None:
                raise ValueError("Could not open image!")

            image_rgb = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            data = pytesseract.image_to_data(image_rgb, output_type=Output.DICT)

            n_boxes = len(data['level'])
            extracted_data = []

            for i in range(n_boxes):
                try:
                    if not data['text'][i].strip():
                        continue
                        
                    x = data['left'][i] + data['width'][i] / 2
                    y = data['top'][i] + data['height'][i] / 2
                    text = data['text'][i]
                    confidence = float(data['conf'][i]) / 100.0 if data['conf'][i] != '-1' else 0.0
                    bbox = [
                        (data['left'][i], data['top'][i]),
                        (data['left'][i] + data['width'][i], data['top'][i]),
                        (data['left'][i] + data['width'][i], data['top'][i] + data['height'][i]),
                        (data['left'][i], data['top'][i] + data['height'][i]),
                    ]
                    extracted_data.append({
                        'x': x, 
                        'y': y, 
                        'text': text.strip(), 
                        'confidence': confidence, 
                        'bbox': bbox
                    })
                except Exception as e:
                    logger.warning(f"Error processing box {i}: {str(e)}")

            return extracted_data

        except Exception as e:
            logger.error(f"Error processing image: {str(e)}")
            raise

    def group_into_rows(self, data, y_threshold=10):
        data_sorted = sorted(data, key=lambda k: k['y'])
        rows = []
        current_row = []
        last_y = None

        for item in data_sorted:
            x = item['x']
            y = item['y']
            text = item['text']
            if last_y is None or abs(y - last_y) > y_threshold:
                if current_row:
                    rows.append(sorted(current_row, key=lambda k: k[0]))
                current_row = [(x, text, item['confidence'])]
                last_y = y
            else:
                current_row.append((x, text, item['confidence']))

        if current_row:
            rows.append(sorted(current_row, key=lambda k: k[0]))

        return [[(text, confidence) for x, text, confidence in row] for row in rows]

    def save_as_xlsx(self, rows, output_xlsx, green_threshold=0.97, yellow_threshold=0.92):
        wb = openpyxl.Workbook()
        ws = wb.active

        for row_index, row in enumerate(rows, start=1):
            for col_index, cell in enumerate(row, start=1):
                text, confidence = cell
                ws.cell(row=row_index, column=col_index, value=text)

                if confidence >= green_threshold:
                    fill_color = '00FF00'
                elif confidence >= yellow_threshold:
                    fill_color = 'FFFF00'
                else:
                    fill_color = 'FF0000'

                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                ws.cell(row=row_index, column=col_index).fill = fill

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_xlsx)

    def draw_bounding_boxes(self, image_path, data, output_path):
        try:
            image = cv2.imread(image_path)
            for item in data:
                bbox = item['bbox']
                if len(bbox) != 4:
                    continue
                
                bbox_np = np.array(bbox, dtype=np.int32)
                bbox_np = bbox_np.reshape((-1, 1, 2))
                
                cv2.polylines(image, [bbox_np], True, (0, 255, 0), 2)
                
                text = item['text']
                confidence = item['confidence']
                label = f"{text} ({confidence:.2f})"
                cv2.putText(image, label, (bbox[0][0], bbox[0][1] - 10),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 0), 2)
            
            cv2.imwrite(output_path, image)
            
        except Exception as e:
            logger.error(f"Error in draw_bounding_boxes: {str(e)}")
            raise

class OCRApp:
    def __init__(self, root):
        # Add is_screenshot initialization
        self.is_screenshot = False
        
        # Simplify to just use current directory
        self.app_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Initialize OCR engines as class attributes
        self.paddle_ocr = PaddleOCREngine(
            model_dir=os.path.join(self.app_dir, 'models')
        )
        self.tesseract_ocr = TesseractOCREngine(
            tesseract_cmd=r'C:\Program Files\Tesseract-OCR\tesseract.exe'
        )
        
        # Initialize the rest of the application
        self.initialize_app(root)

    def initialize_app(self, root):
        self.root = root
        self.root.title("OCR to Excel Tool")
        self.root.geometry("1280x720")
        
        # Use simple path for icon
        icon_path = 'icons/icon.png'
        self.icon_image = tk.PhotoImage(file=icon_path)
        self.root.iconphoto(False, self.icon_image)
        
        self.style = ttk.Style('cosmo')
        self.ocr_engine = tk.StringVar()
        self.ocr_engine.set("PaddleOCR")

        self.green_threshold = tk.IntVar(value=97)
        self.yellow_threshold = tk.IntVar(value=92)
        self.output_directory = None
        
        self.setup_ui()

    def setup_ui(self):
        # Main frame
        self.main_frame = ttk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Center frame for initial view
        self.center_frame = ttk.Frame(self.main_frame)
        self.center_frame.pack(expand=True)

        # Logo
        logo_img = Image.open("icons/icon.png")
        logo_img = logo_img.resize((100, 100), Image.LANCZOS)
        self.logo_photo = ImageTk.PhotoImage(logo_img)
        self.logo_label = ttk.Label(self.center_frame, image=self.logo_photo)
        self.logo_label.pack(pady=(20, 10))

        # Output Directory Selection
        folder_icon = Image.open("icons/folder.png")
        folder_icon = folder_icon.resize((20, 20), Image.LANCZOS)
        self.folder_icon_photo = ImageTk.PhotoImage(folder_icon)
        output_dir_button = ttk.Button(self.center_frame, text="Select Output Directory", 
                                       image=self.folder_icon_photo, compound=tk.LEFT,
                                       command=self.select_output_directory)
        output_dir_button.pack(pady=(10, 5))

        # OCR Engine Selection
        ocr_label = ttk.Label(self.center_frame, text="Select OCR Engine:")
        ocr_label.pack(pady=(10, 5))
        ocr_dropdown = ttk.Combobox(self.center_frame, textvariable=self.ocr_engine, state="readonly", width=30)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract')
        ocr_dropdown.pack(pady=(0, 20))

        # Confidence Thresholds
        thresholds_frame = ttk.Frame(self.center_frame)
        thresholds_frame.pack(pady=(10, 20))

        options = [str(i) for i in range(80, 101)]  # Values from 80 to 100

        green_label = ttk.Label(thresholds_frame, text="High Confidence Threshold (Green) (> %):")
        green_label.grid(row=0, column=0, padx=5, pady=5, sticky='e')
        green_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.green_threshold, values=options, state='readonly', width=5)
        green_dropdown.grid(row=0, column=1, padx=5, pady=5)
        green_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        yellow_label = ttk.Label(thresholds_frame, text="Medium Confidence Threshold (Yellow) (> %):")
        yellow_label.grid(row=1, column=0, padx=5, pady=5, sticky='e')
        yellow_dropdown = ttk.Combobox(thresholds_frame, textvariable=self.yellow_threshold, values=options, state='readonly', width=5)
        yellow_dropdown.grid(row=1, column=1, padx=5, pady=5)
        yellow_dropdown.bind('<<ComboboxSelected>>', self.update_thresholds)

        # Upload Button
        upload_icon = Image.open("icons/upload.png")
        upload_icon = upload_icon.resize((20, 20), Image.LANCZOS)
        self.upload_icon_photo = ImageTk.PhotoImage(upload_icon)
        self.upload_button = ttk.Button(self.center_frame, text="Upload Image", 
                                        image=self.upload_icon_photo, compound=tk.LEFT,
                                        command=self.select_image, width=20)
        self.upload_button.pack(pady=(0, 10))

        # Screenshot Button
        screenshot_icon = Image.open("icons/screenshot.png")
        screenshot_icon = screenshot_icon.resize((20, 20), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        self.screenshot_button = ttk.Button(self.center_frame, text="Screenshot", 
                                            image=self.screenshot_icon_photo, compound=tk.LEFT,
                                            command=self.take_screenshot, width=20)
        self.screenshot_button.pack(pady=(0, 20))

        # Status Label
        self.status_label = ttk.Label(self.center_frame, text="")
        self.status_label.pack(pady=(0, 10))

        # Progress bar (hidden initially)
        self.progress_bar = ttk.Progressbar(self.center_frame, mode='indeterminate')
        self.progress_bar.pack(pady=(0, 10))
        self.progress_bar.pack_forget()

        # Prepare frames for results
        self.top_frame = ttk.Frame(self.main_frame)
        self.left_frame = ttk.Frame(self.main_frame)
        self.middle_frame = ttk.Frame(self.main_frame)
        self.right_frame = ttk.Frame(self.main_frame)  # For the sidebar

    def select_output_directory(self):
        directory = filedialog.askdirectory()
        if directory:
            self.output_directory = directory
            self.status_label.config(text=f"Output directory set to: {self.output_directory}")

    def update_thresholds(self, event=None):
        # Ensure that green_threshold is always greater than yellow_threshold
        if self.green_threshold.get() < self.yellow_threshold.get():
            self.yellow_threshold.set(self.green_threshold.get())

    def reset_ui(self):
        # Hide the top, left, middle, and right frames
        self.top_frame.pack_forget()
        self.left_frame.pack_forget()
        self.middle_frame.pack_forget()
        self.right_frame.pack_forget()

        # Clear images from frames
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Show the center frame
        self.center_frame.pack(expand=True)

        # Reset status label
        self.status_label.config(text="")

        # Remove progress bar
        self.progress_bar.pack_forget()

    def select_image(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Image files", ("*.png", "*.jpg", "*.jpeg", "*.bmp", "*.tiff"))]
        )
        if file_path:
            self.is_screenshot = False  # Set to False for uploaded images
            self.current_image_path = file_path  # Store the current image path
            self.reset_ui()
            self.process_image(file_path)

    def take_screenshot(self):
        self.root.withdraw()
        image = self.capture_screenshot()
        
        if image:
            self.root.deiconify()
            self.is_screenshot = True  # Set to True for screenshots
            
            if self.output_directory:
                output_dir = self.output_directory
            else:
                output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            
            os.makedirs(output_dir, exist_ok=True)
            base_filename = "screenshot"
            date_string = time.strftime("%Y-%m-%d_%H-%M-%S")
            screenshot_path = os.path.join(output_dir, base_filename + "_" + date_string + ".png")
            image.save(screenshot_path)
            self.current_image_path = screenshot_path  # Store the current image path
            self.reset_ui()
            self.process_image(screenshot_path)

    def capture_screenshot(self):
        
        from PIL import ImageGrab        
        try:
            # Store current window state
            window_state = self.root.state()
            
            # Minimize the window
            self.root.withdraw()

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
            
            # Simulate Windows + Shift + S keystroke
            keyboard.press_and_release('windows+shift+s')
            start_time = time.time()
            # Wait for 10 seconds
            timeout = 30

            while True:
                # Try to get the screenshot from clipboard
                screenshot = ImageGrab.grabclipboard() 
                if screenshot is not None:
                    return screenshot
                else:
                    # Restore window before showing error
                    if time.time() - start_time > timeout:
                        self.root.deiconify()
                        messagebox.showerror("Screenshot Error", 
                        "No screenshot was taken within the time limit (30 seconds).\n"
                        "Please try again and make sure to complete the screenshot within the time limit.")
                        raise TimeoutError("Retry taking screenshot within 30 seconds")
                    time.sleep(2)
        except TimeoutError as e:
            raise e
        except Exception as e:
            # Always restore window in case of any error
            self.root.deiconify()
            messagebox.showerror("Error", f"An error occurred: {str(e)}")
            raise Exception(f"Error capturing screenshot: {str(e)}")

    def process_image(self, file_path):
        # Start a new thread for processing
        processing_thread = threading.Thread(target=self._process_image_thread, args=(file_path,))
        processing_thread.start()

    def _process_image_thread(self, file_path):
        try:
            ocr_engine = self.ocr_engine.get()

            # Show progress bar in the center frame
            self.progress_bar.pack(pady=(0, 10))
            self.progress_bar.start()

            if ocr_engine == "PaddleOCR":
                self.process_with_paddleocr(file_path)
            elif ocr_engine == "Tesseract":
                self.process_with_tesseract(file_path)
            else:
                raise ValueError("Please select an OCR engine.")
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")
        finally:
            # Hide progress bar
            self.progress_bar.stop()
            self.progress_bar.pack_forget()

    def process_with_paddleocr(self, file_path):
        try:
            data = self.paddle_ocr.process_image(file_path)
            rows = self.paddle_ocr.group_into_rows(data)
            
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    output_dir = os.path.dirname(file_path)
                
            os.makedirs(output_dir, exist_ok=True)
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            self.paddle_ocr.save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)
            self.paddle_ocr.draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")

    def process_with_tesseract(self, file_path):
        try:
            data = self.tesseract_ocr.process_image(file_path)

            if not data:
                raise ValueError("No data extracted from image.")

            rows = self.tesseract_ocr.group_into_rows(data)

            if not rows:
                raise ValueError("No rows extracted from data.")

            # Determine the output directory
            if self.output_directory:
                output_dir = self.output_directory
            else:
                if self.is_screenshot:
                    output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
                else:
                    output_dir = os.path.dirname(file_path)
                
            os.makedirs(output_dir, exist_ok=True)
            base_filename = os.path.splitext(os.path.basename(file_path))[0]
            output_xlsx = os.path.join(output_dir, base_filename + "_output.xlsx")
            output_image_path = os.path.join(output_dir, base_filename + "_output_image.jpg")

            green_thresh = self.green_threshold.get() / 100.0
            yellow_thresh = self.yellow_threshold.get() / 100.0

            self.tesseract_ocr.save_as_xlsx(rows, output_xlsx, green_thresh, yellow_thresh)
            self.tesseract_ocr.draw_bounding_boxes(file_path, data, output_image_path)

            self.status_label.config(text=f"Excel file saved: {output_xlsx}")
            self.display_results(output_image_path, output_xlsx)
        except Exception as e:
            self.status_label.config(text=f"Error: {str(e)}\nPlease try a different image or OCR engine.")

    def display_results(self, image_path, excel_path):
        self.reorganize_layout()

        # Clear previous images
        for widget in self.left_frame.winfo_children():
            widget.destroy()
        for widget in self.middle_frame.winfo_children():
            widget.destroy()
        for widget in self.right_frame.winfo_children():
            widget.destroy()

        # Display image with bounding boxes
        self.display_image(image_path, self.left_frame)

        # Display Excel image
        excel_image_path = os.path.splitext(excel_path)[0] + "_excel_image.png"
        self.generate_excel_image(excel_path, excel_image_path)

        # Add padding to the middle frame
        padding_frame = ttk.Frame(self.middle_frame, padding=20)
        padding_frame.pack(fill=tk.BOTH, expand=True)
        self.display_image(excel_image_path, padding_frame)

        # Setup the sidebar
        self.setup_sidebar()

    def reorganize_layout(self):
        # Hide the center frame
        self.center_frame.pack_forget()

        # Clear the top_frame
        for widget in self.top_frame.winfo_children():
            widget.destroy()

        # Set up top frame
        self.top_frame.pack(side=tk.TOP, fill=tk.X, pady=(10, 0))

        # Create a frame inside top_frame to center widgets
        top_inner_frame = ttk.Frame(self.top_frame)
        top_inner_frame.pack(anchor='center')

        # Add small logo
        small_logo_img = Image.open("icons/icon.png")
        small_logo_img = small_logo_img.resize((50, 50), Image.LANCZOS)
        self.small_logo_photo = ImageTk.PhotoImage(small_logo_img)
        logo_label = ttk.Label(top_inner_frame, image=self.small_logo_photo)
        logo_label.pack(side=tk.LEFT, padx=(10, 5))

        ocr_label = ttk.Label(top_inner_frame, text="Select OCR Engine:")
        ocr_label.pack(side=tk.LEFT, padx=(10, 5))
        ocr_dropdown = ttk.Combobox(top_inner_frame, textvariable=self.ocr_engine, state="readonly", width=20)
        ocr_dropdown['values'] = ('PaddleOCR', 'Tesseract')
        ocr_dropdown.pack(side=tk.LEFT, padx=(0, 10))
        upload_button = ttk.Button(top_inner_frame, text="Upload Image", command=self.select_image)
        upload_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Screenshot Button with Icon
        screenshot_icon = Image.open(os.path.join('icons', 'screenshoticon.png'))
        screenshot_icon = screenshot_icon.resize((30, 30), Image.LANCZOS)
        self.screenshot_icon_photo = ImageTk.PhotoImage(screenshot_icon)
        screenshot_button = ttk.Button(top_inner_frame, image=self.screenshot_icon_photo, command=self.take_screenshot)
        screenshot_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Open Folder Button with Icon
        folder_icon = Image.open(os.path.join('icons', 'folder.png'))
        folder_icon = folder_icon.resize((30, 30), Image.LANCZOS)
        self.folder_icon_photo = ImageTk.PhotoImage(folder_icon)
        folder_button = ttk.Button(top_inner_frame, image=self.folder_icon_photo, command=self.open_output_folder)
        folder_button.pack(side=tk.LEFT, padx=(0, 10))

        # Add Home Button with Icon
        home_icon = Image.open(os.path.join('icons', 'home.png'))
        home_icon = home_icon.resize((30, 30), Image.LANCZOS)
        self.home_icon_photo = ImageTk.PhotoImage(home_icon)
        home_button = ttk.Button(top_inner_frame, image=self.home_icon_photo, command=self.reset_ui)
        home_button.pack(side=tk.LEFT, padx=(0, 10))

        # Set up frames
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.middle_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)

    def display_image(self, image_path, panel):
        image = Image.open(image_path)

        # Create a canvas to display the image
        canvas = tk.Canvas(panel, bg='white')
        canvas.pack(fill=tk.BOTH, expand=True)

        # Function to resize the image when the panel size changes
        def resize_image(event):
            # Get the size of the canvas
            canvas_width = event.width
            canvas_height = event.height

            # Calculate the scaling factor
            width_ratio = canvas_width / image.width
            height_ratio = canvas_height / image.height
            scale_factor = min(width_ratio, height_ratio, 1.0)  # Do not upscale images

            new_width = int(image.width * scale_factor)
            new_height = int(image.height * scale_factor)

            resized_image = image.resize((new_width, new_height), Image.LANCZOS)
            photo = ImageTk.PhotoImage(resized_image)

            canvas.delete("all")
            canvas.create_image(canvas_width/2, canvas_height/2, image=photo, anchor='center')
            canvas.image = photo  # Keep a reference

        # Bind the resize event to the function
        canvas.bind("<Configure>", resize_image)

    def generate_excel_image(self, excel_path, output_image_path):
        from openpyxl import load_workbook
        from PIL import Image, ImageDraw, ImageFont

        wb = load_workbook(excel_path)
        ws = wb.active

        # Increase cell size and font size
        cell_width = 90
        cell_height = 30
        font_size = 20
        border_size = 40  # Increased border size

        max_col = ws.max_column
        max_row = ws.max_row
        image_width = cell_width * max_col + 2 * border_size
        image_height = cell_height * max_row + 2 * border_size

        image = Image.new('RGB', (image_width, image_height), 'white')
        draw = ImageDraw.Draw(image)
        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except:
            font = ImageFont.load_default()

        for row in ws.iter_rows():
            for cell in row:
                col_idx = cell.column - 1
                row_idx = cell.row - 1
                x1 = col_idx * cell_width + border_size
                y1 = row_idx * cell_height + border_size
                x2 = x1 + cell_width
                y2 = y1 + cell_height

                # Draw cell background
                fill_color = 'FFFFFF'  # Default fill
                if cell.fill and cell.fill.fgColor:
                    if cell.fill.fgColor.type == 'rgb':
                        fill_color = cell.fill.fgColor.rgb[-6:]
                    elif cell.fill.fgColor.type == 'indexed':
                        fill_color = 'FFFFFF'  # Handle indexed colors as white

                draw.rectangle([x1, y1, x2, y2], fill=f'#{fill_color}', outline='black')

                # Draw cell text
                text = str(cell.value) if cell.value is not None else ''
                bbox = font.getbbox(text)
                text_width = bbox[2] - bbox[0]
                text_height = bbox[3] - bbox[1]
                text_x = x1 + (cell_width - text_width) / 2
                text_y = y1 + (cell_height - text_height) / 2
                draw.text((text_x, text_y), text, fill='black', font=font)

        image.save(output_image_path)

    def setup_sidebar(self):
        # Sidebar content
        sidebar_frame = ttk.Frame(self.right_frame, padding=10)
        sidebar_frame.pack(fill=tk.Y, expand=False)

        # Accuracy Box
        accuracy_label = ttk.Label(sidebar_frame, text="Accuracy Thresholds", font=('Helvetica', 14, 'bold'))
        accuracy_label.pack(pady=(0, 10))

        # Icons for accuracy levels
        green_icon = Image.open(os.path.join('icons', 'green_circle.png'))
        green_icon = green_icon.resize((20, 20), Image.LANCZOS)
        self.green_icon_photo = ImageTk.PhotoImage(green_icon)

        yellow_icon = Image.open(os.path.join('icons', 'yellow_circle.png'))
        yellow_icon = yellow_icon.resize((20, 20), Image.LANCZOS)
        self.yellow_icon_photo = ImageTk.PhotoImage(yellow_icon)

        red_icon = Image.open(os.path.join('icons', 'red_circle.png'))
        red_icon = red_icon.resize((20, 20), Image.LANCZOS)
        self.red_icon_photo = ImageTk.PhotoImage(red_icon)

        # Green Threshold Label
        green_frame = ttk.Frame(sidebar_frame)
        green_frame.pack(pady=5, anchor='w')
        green_label_icon = ttk.Label(green_frame, image=self.green_icon_photo)
        green_label_icon.pack(side=tk.LEFT)
        green_label_text = ttk.Label(green_frame, text=f"High Confidence (> {self.green_threshold.get()}%)")
        green_label_text.pack(side=tk.LEFT, padx=5)

        # Yellow Threshold Label
        yellow_frame = ttk.Frame(sidebar_frame)
        yellow_frame.pack(pady=5, anchor='w')
        yellow_label_icon = ttk.Label(yellow_frame, image=self.yellow_icon_photo)
        yellow_label_icon.pack(side=tk.LEFT)
        yellow_label_text = ttk.Label(yellow_frame, text=f"Medium Confidence (> {self.yellow_threshold.get()}%)")
        yellow_label_text.pack(side=tk.LEFT, padx=5)

        # Red Threshold Label
        red_frame = ttk.Frame(sidebar_frame)
        red_frame.pack(pady=5, anchor='w')
        red_label_icon = ttk.Label(red_frame, image=self.red_icon_photo)
        red_label_icon.pack(side=tk.LEFT)
        red_label_text = ttk.Label(red_frame, text=f"Low Confidence (≤ {self.yellow_threshold.get()}%)")
        red_label_text.pack(side=tk.LEFT, padx=5)

        # Divider
        separator = ttk.Separator(sidebar_frame, orient='horizontal')
        separator.pack(fill='x', pady=10)

        # Disclaimer Box
        disclaimer_frame = ttk.Frame(sidebar_frame)
        disclaimer_frame.pack(pady=(10, 10))

        # Info Icon
        info_icon = Image.open(os.path.join('icons', 'info.png'))
        info_icon = info_icon.resize((20, 20), Image.LANCZOS)
        self.info_icon_photo = ImageTk.PhotoImage(info_icon)

        disclaimer_title_frame = ttk.Frame(disclaimer_frame)
        disclaimer_title_frame.pack(anchor='w')

        disclaimer_icon_label = ttk.Label(disclaimer_title_frame, image=self.info_icon_photo)
        disclaimer_icon_label.pack(side=tk.LEFT)

        disclaimer_label = ttk.Label(disclaimer_title_frame, text="Disclaimer", font=('Helvetica', 14, 'bold'))
        disclaimer_label.pack(side=tk.LEFT, padx=5)

        # Enhanced Disclaimer Text
        disclaimer_text = (
            "Note: When the input image contains a table with missing data in some rows or columns, "
            "the extracted data may not align perfectly in the output Excel file. "
            "Empty cells might cause subsequent data to shift positions, resulting in misaligned columns. "
            "Please review the Excel output carefully and adjust as needed."
        )
        disclaimer_message = ttk.Label(disclaimer_frame, text=disclaimer_text, wraplength=250, justify='left')
        disclaimer_message.pack(pady=5)

    def open_output_folder(self):
        """Open the output directory in the system's file explorer"""
        if self.output_directory:
            output_dir = self.output_directory
        else:
            # If no output directory was selected, use Desktop for screenshots or the image's directory
            if hasattr(self, 'is_screenshot') and self.is_screenshot:
                output_dir = os.path.join(os.path.expanduser("~"), "Desktop")
            else:
                output_dir = os.path.dirname(self.current_image_path) if hasattr(self, 'current_image_path') else None

        if output_dir and os.path.exists(output_dir):
            if sys.platform == 'darwin':  # macOS
                subprocess.run(['open', output_dir])
            elif sys.platform == 'win32':  # Windows
                subprocess.run(['explorer', output_dir])
            else:  # Linux
                subprocess.run(['xdg-open', output_dir])
        else:
            self.status_label.config(text="Output directory not found")

if __name__ == "__main__":
    root = ttk.Window(themename="cosmo")
    app = OCRApp(root)
    root.mainloop()